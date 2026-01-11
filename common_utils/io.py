import glob
import json
import logging
from pathlib import Path
from typing import List, Union, Generator, Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter  # 核心修复：支持无限列名
import tqdm

# 设置简单的日志打印，避免数据静默丢失
logger = logging.getLogger(__name__)
if not logger.handlers:
    logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)

def read_jsonl(path_patterns: Union[str, List[str]], is_root: bool = False, ignore_errors: bool = False) -> Generator[Any, None, None]:
    """
    读取 JSONL 文件。
    :param ignore_errors: 是否忽略解析错误的行（默认为 False，会打印 Warning）
    """
    if isinstance(path_patterns, str):
        path_patterns = [path_patterns]
    
    total_files = 0
    for path_pattern in path_patterns:
        # glob 排序确保处理顺序可复现
        path_list = sorted(glob.glob(path_pattern, recursive=True))

        if not path_list:
            logger.warning(f"路径模式未匹配到任何文件: {path_pattern}")
            continue

        total_files += len(path_list)
        for path in path_list:
            line_count = 0
            # errors="replace" 防止遇到非 utf-8 字符直接崩溃
            with open(path, encoding="utf-8", errors="replace") as reader:
                for line_num, line in enumerate(reader, 1):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        line_json = json.loads(line)
                        if isinstance(line_json, dict) and is_root:
                            line_json["##FILEPATH##"] = path
                        line_count += 1
                        yield line_json
                    except json.JSONDecodeError as e:
                        if not ignore_errors:
                            logger.warning(f"JSON解析失败 [{path}:Line {line_num}]: {e}")
                        continue
                    except Exception as e:
                        if not ignore_errors:
                            logger.warning(f"读取异常 [{path}:Line {line_num}]: {e}")
                        continue
            # 读完一个文件打印统计，确认文件不是空的
            logger.debug(f"文件读取完成: {path} (行数: {line_count})")
    if total_files == 0:
        logger.error("未找到任何符合条件的输入文件！生成器将为空。")

def read_json(path_patterns: Union[str, List[str]], ignore_errors: bool = False) -> Generator[Any, None, None]:
    """
    读取标准 JSON 文件（注意：大文件可能导致 OOM）。
    """
    if isinstance(path_patterns, str):
        path_patterns = [path_patterns]
    
    for path_pattern in path_patterns:
        path_list = sorted(glob.glob(path_pattern, recursive=True))
        for path in path_list:
            with open(path, encoding="utf-8", errors="replace") as reader:
                try:
                    yield json.load(reader)
                except Exception as e:
                    if not ignore_errors:
                        logger.warning(f"JSON文件加载失败 [{path}]: {e}")
                    continue

def save_jsonl(samples, result_save_path: str, overwrite: bool = False) -> None:
    """
    保存为 JSONL 格式。
    """
    p = Path(result_save_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    
    if p.exists() and not overwrite:
        logger.info(f"文件[{result_save_path}]已存在，跳过写入 (overwrite=False)。")
        return
    
    try:
        count = 0
        with open(result_save_path, "w", encoding="utf-8") as writer:
            for sample in samples:
                # 确保写入的是单行 JSON
                writer.write(json.dumps(sample, ensure_ascii=False) + "\n")
                count += 1
        logger.info(f"已保存: {result_save_path} (共 {count} 条)")
    except Exception as e:
        logger.error(f"写入文件失败 [{result_save_path}]: {e}")
        raise

def save_data_to_excel_merge(samples, file_path, key, merge=True):
    """
    生成表格，并根据 key 字段对应的列表进行单元格合并。
    修复了列名超过 Z (26列) 报错的问题。
    """
    samples = list(samples)
    if not samples:
        logger.warning(f"没有数据需要保存到 Excel: {file_path}")
        return
    
    # 基础校验
    first = samples[0]
    if key not in first:
        raise ValueError(f"主键 '{key}' 不在数据字段中")
    if not isinstance(first[key], list) or len(first[key]) == 0 or not isinstance(first[key][0], dict):
        raise ValueError(f"字段 '{key}' 必须是包含字典的非空列表 (List[Dict])，用于展开合并。")

    # 构建列名
    column_names = [k for k in first if k != key]
    # 展开内部 list 的 keys
    for k1 in first[key][0]:
        column_names.append(f"{key}::{k1}")

    wb = openpyxl.Workbook()
    ws = wb.active
    font_bold = Font(bold=True)
    
    # 构建列名映射：Name -> "A", "B", ... "AA", "AB"
    # 使用 enumerate(..., 1) 配合 get_column_letter
    col_mapping = {}
    for idx, name in enumerate(column_names, 1):
        col_letter = get_column_letter(idx)
        col_mapping[name] = col_letter
        # 写入表头
        cell = ws[f"{col_letter}1"]
        cell.value = name
        cell.font = font_bold

    row_start = 2  # 第一行是标题
    for s in tqdm.tqdm(samples, desc=f"Writing to {Path(file_path).name}"):
        nrow = len(s[key])
        row_end = row_start + nrow - 1
        
        # 1. 填入外部公共字段 (需合并)
        for k in s:
            if k == key: continue
            if k not in col_mapping: continue # 防御性编程
            
            col = col_mapping[k]
            
            # 合并单元格
            if merge and nrow > 1:
                ws.merge_cells(f"{col}{row_start}:{col}{row_end}")
            
            val = s[k]
            # 如果不是基本类型，转 JSON 字符串，防止 Excel 写入报错
            if not isinstance(val, (str, int, float, bool, type(None))):
                val = json.dumps(val, ensure_ascii=False)
            
            # 填入起始行
            ws[f"{col}{row_start}"] = val

        # 2. 填入内部展开字段 (逐行填入)
        for ii, inner_item in enumerate(s[key]):
            current_row = row_start + ii
            for kk, vv in inner_item.items():
                col_name = f"{key}::{kk}"
                if col_name in col_mapping:
                    col = col_mapping[col_name]
                    
                    if not isinstance(vv, (str, int, float, bool, type(None))):
                        vv = json.dumps(vv, ensure_ascii=False)
                    
                    ws[f"{col}{current_row}"] = vv
        
        row_start += nrow

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    logger.info(f"✅ Excel saved: {file_path} - Records: {len(samples)}")